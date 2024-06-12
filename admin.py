from telebot import types, TeleBot
from DBManager import DBManager
from Model import Model
from models import *
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import copy, os, logging, json, uuid

admin_markup = types.InlineKeyboardMarkup()
admin_markup.add(types.InlineKeyboardButton("Тесты:", callback_data="None"), row_width=1)
admin_markup.add(
    types.InlineKeyboardButton("Выгрузить", callback_data=json.dumps({"c": "get-tests"})),
    types.InlineKeyboardButton("Загрузить", callback_data=json.dumps({"c": "update-tests"})),
    row_width=2
)
admin_markup.add(types.InlineKeyboardButton("Теория:", callback_data="None"), row_width=1)
admin_markup.add(
    types.InlineKeyboardButton("Выгрузить", callback_data=json.dumps({"c": "get-theory"})),
    types.InlineKeyboardButton("Загрузить", callback_data=json.dumps({"c": "update-theory"})),
    row_width=2
)
admin_markup.add(types.InlineKeyboardButton("Загрузить файл", callback_data=json.dumps({"c": "update-file"})))
admin_markup.add(types.InlineKeyboardButton("База данных пользователей", callback_data=json.dumps({"c": "get-users-table"})))

def get_cell(cells, x, y):
    return str(cells.get(x, {y: ""}).get(y, ""))

def rand_str():
    return str(uuid.uuid4())[:5]
def get_cells(sheet):
    cells = {}
    for y, column in enumerate(sheet):
        for x, cell in enumerate(column):
            if cells.get(x+1) is None:
                cells[x+1] = {}
            if cell.value is None:
                cells[x+1][y+1] = None
            else:
                cells[x+1][y+1] = str(cell.value)
    return cells

def get_files(cells, x, y) -> list[str]:
    ans = []
    for i in range(1000):
        if not cells.get(x, {y+i: None}).get(y+i):
            continue
        ans.append(cells[x][y+i])
    return ans

def set_files(sheet, files, x, y):
    for i, file in enumerate(files):
        sheet.cell(y+i, x, file)

def replace_none_to_space(a):
    if isinstance(a, Model):
        d = a.__dict__
    elif isinstance(a, dict):
        d = a
    else:
        return a
    for key, value in d.items():
        if value is None or isinstance(value, str) and value.strip() == "None":
            d[key] = " "
    return a

def read_excel_theory(filepath: str):
    wb = load_workbook(filepath)
    data = {}
    for sheet in wb:
        cells = get_cells(sheet)
        data[sheet.title] = {
            "name": get_cell(cells, 2, 2), 
            "description": get_cell(cells, 2, 3), 
            "files": get_files(cells, 2, 4), 
            "paragraphs": []
        }
        for i in range(100):
            if not get_cell(cells, 3+i, 2):
                continue
            data[sheet.title]["paragraphs"].append({
                "name": get_cell(cells, 3+i, 2),
                "description": get_cell(cells, 3+i, 3),
                "files": get_files(cells, 3+i, 4),
            })
    return [value for _, value in data.items()]

def read_excel_tests(filepath: str):
    wb = load_workbook(filepath)
    data = {}
    for sheet in wb:
        cells = get_cells(sheet)
        data[sheet.title] = {
            "name": get_cell(cells, 2, 2), 
            "description": get_cell(cells, 2, 3), 
            "files": get_files(cells, 2, 5), 
            "tasks": []
        }
        for i in range(100):
            if not(get_cell(cells, 3+i, 1).upper().strip() in ["ВОПРОС", "QUESTION"]):
                continue
            task = {
                "name": get_cell(cells, 3+i, 2),
                "description": get_cell(cells, 3+i, 3),
                "right_answer": get_cell(cells, 3+i, 4),
                "files": get_files(cells, 3+i, 5),
                "answers": []
            }
            for j in range(1, 10):
                if not(get_cell(cells, 3+i+j, 1).upper().strip() in ["ОТВЕТ", "ANSWER"]):
                    break
                task["answers"].append({
                    "name": get_cell(cells, 3+i+j, 2),
                    "description": get_cell(cells, 3+i+j, 3),
                    "files": get_files(cells, 3+i+j, 5),
                })
            data[sheet.title]["tasks"].append(task)
    wb.close()
    return [value for _, value in data.items()]

def get_users_table(bot: TeleBot, callback: types.CallbackQuery, user: Model, db_manager: DBManager):
    wb = Workbook()
    sheet = wb.active
    sheet.cell(1, 1, "Телеграм-id")
    sheet.cell(1, 2, "Телеграм имя")
    sheet.cell(1, 3, "Имя")
    sheet.cell(1, 4, "Класс")
    sheet.cell(1, 5, "Первое сообщение")
    sheet.cell(1, 6, "Состоние")
    sheet.cell(1, 7, "Пройденные тесты")
    sheet.cell(1, 8, "Пройденные теории")
    sheet.cell(1, 9, "Текущий тест")
    for i in range(1, 10):
        sheet.column_dimensions[get_column_letter(i)].width = max(len(str(sheet.cell(1, i).value)), 10)
    for i,  user in enumerate(db_manager.find_data(UserModel)):
        user = replace_none_to_space(user)
        sheet.cell(2+i, 1, user.telegram_id)
        sheet.cell(2+i, 2, user.user_name)
        sheet.cell(2+i, 3, user.full_name)
        sheet.cell(2+i, 4, user.grade)
        sheet.cell(2+i, 5, user.first_message)
        sheet.cell(2+i, 6, user.state)
        sheet.cell(2+i, 7, str(user.accepted_tests))
        sheet.cell(2+i, 8, str(user.accepted_theory))
        sheet.cell(2+i, 9, str(user.current_test))
    os.makedirs("./tmp/files", exist_ok=True)
    filepath = "./tmp/files/users.xlsx"
    wb.save(filepath)
    wb.close()
    bot.send_document(user.telegram_id, open(filepath, "rb"))

def update_tests_cb(bot: TeleBot, callback: types.CallbackQuery, user: Model, db_manager: DBManager):
    user.state = "update_tests"
    db_manager.save_data(user)
    bot.send_message(user.telegram_id, "Отправьте таблицу с тестами")

def update_tests(bot: TeleBot, filepath: str, message: types.Message, user: Model, db_manager: DBManager):
    user.state = "menu"
    db_manager.save_data(user)
    db_manager.delete_data(TestModel, condition="true")
    tests = read_excel_tests(filepath)
    for test in tests:
        print(test)
        test_model = copy.deepcopy(TestModel)
        test_model.name = test["name"]
        test_model.description = test["description"]
        test_model.files = test["files"]
        test_model.tasks = test["tasks"]
        print("save test")
        db_manager.save_data(test_model)
    os.remove(filepath)
    bot.send_message(user.telegram_id, "Отправьте /menu, чтобы появилась клавиатура")
    return True

def get_tests_cb(bot: TeleBot, callback: types.CallbackQuery, user: Model, db_manager: DBManager):
    tests = db_manager.find_data(TestModel)
    os.makedirs("./tmp/files/", exist_ok=True)
    filepath = "./tmp/files/tests.xlsx"
    wb = Workbook()
    if not tests:
        sheet = wb.active
        sheet.cell(1, 1, "1")
        sheet.cell(2, 1, "Название")
        sheet.cell(3, 1, "Описание")
        sheet.cell(4, 1, "Правильный ответ")
        sheet.cell(5, 1, "Файлы")

        sheet.cell(1, 2, "Тест")
        sheet.cell(2, 2, "Название теста")
        sheet.cell(3, 2, "Описание теста")
        sheet.cell(4, 2, "-")
        sheet.cell(5, 2, f"Файл теста {rand_str()}.jpg")
        sheet.cell(6, 2, f"{rand_str()}.jpg")
        sheet.cell(7, 2, f"{rand_str()}.jpg")

        current_column = 3
        for i in range(3):
            sheet.cell(1, current_column, "Вопрос")
            sheet.cell(2, current_column, f"Название вопроса")
            sheet.cell(3, current_column, f"Описание вопроса")
            sheet.cell(4, current_column, f"1")
            sheet.cell(5, current_column, f"Файл вопроса {rand_str()}.jpg")
            sheet.cell(6, current_column, f"{rand_str()}.jpg")
            sheet.cell(7, current_column, f"{rand_str()}.jpg")
            for _ in range(3):
                current_column += 1
                sheet.cell(1, current_column, "Ответ")
                sheet.cell(2, current_column, "Название ответа")
                sheet.cell(3, current_column, "Описание ответа")
                sheet.cell(4, current_column, f"-")
                sheet.cell(5, current_column, f"Файл ответа {rand_str()}.jpg")
                sheet.cell(6, current_column, f"{rand_str()}.jpg")
                sheet.cell(7, current_column, f"{rand_str()}.jpg")
            current_column += 2
        bot.send_message(user.telegram_id, "Тесты отсутствуют. Вот пример заполнения таблицы:")
    else:
        for i, test in enumerate(tests):
            test = replace_none_to_space(test)
            sheet_name = copy.deepcopy(test.name)
            sheet_name = "".join(filter(lambda x: x in "йцукенгшщзхфывапролдячсмитьбюЙЦУКЕНГШЩЗХЪъФЫВАПРОЛДЖЭЯЧСМИТЬБЮqwertyuiopasdfghjklzxcvbnmQWERTYUIOPASDFGHJKLXCVBNM",sheet_name))
            if not sheet_name:
                sheet_name = rand_str()
            sheet = wb.create_sheet(sheet_name)
            sheet.cell(1, 1, str(i+1))
            sheet.cell(1, 2, "Тест")
            sheet.cell(2, 1, "Название")
            sheet.cell(2, 2, test.name)
            sheet.cell(3, 1, "Описание")
            sheet.cell(3, 2, test.description)
            sheet.cell(4, 1, "Правильный ответ")
            sheet.cell(5, 1, "Файлы")
            set_files(sheet, test.files, 2, 5)
            last_task_cell = 3
            for task in test.tasks:
                task = replace_none_to_space(task)
                sheet.cell(1, last_task_cell, "Вопрос")
                sheet.cell(2, last_task_cell, task["name"])
                sheet.cell(3, last_task_cell, task["description"])
                sheet.cell(4, last_task_cell, task["right_answer"])
                set_files(sheet, task["files"], last_task_cell, 5)
                for answer in task["answers"]:
                    answer = replace_none_to_space(answer)
                    last_task_cell += 1
                    sheet.cell(1, last_task_cell, "Ответ")
                    sheet.cell(2, last_task_cell, answer["name"])
                    sheet.cell(3, last_task_cell, answer["description"])
                    sheet.cell(4, last_task_cell, "")
                    set_files(sheet, answer["files"], last_task_cell, 5)
                last_task_cell += 2
        wb.remove(wb[wb.worksheets[0].title])
    os.makedirs("./tmp/files/", exist_ok=True)
    filepath = "./tmp/files/tests.xlsx"
    wb.save(filepath)
    wb.close()
    bot.send_document(user.telegram_id, open(filepath, "rb"))

def update_theory_cb(bot: TeleBot, callback: types.CallbackQuery, user: Model, db_manager: DBManager):
    user.state = "update_theory"
    db_manager.save_data(user)
    bot.send_message(user.telegram_id, "Отправьте таблицу с теорией")

def update_theory(bot: TeleBot, filepath: str, callback: types.CallbackQuery, user: Model, db_manager: DBManager):
    user.state = "menu"
    db_manager.save_data(user)
    bot.send_message(user.telegram_id, "Отправьте /menu, чтобы появилась клавиатура")
    theorys = read_excel_theory(filepath)
    print(theorys)
    db_manager.delete_data(TheoryModel, "true")
    for theory in theorys:
        theory_model = copy.deepcopy(TheoryModel)
        theory_model.name = theory["name"]
        theory_model.description = theory["description"]
        theory_model.files = theory["files"]
        theory_model.paragraphs = theory["paragraphs"]
        db_manager.save_data(theory_model)
    return True

def get_theory_cb(bot: TeleBot, callback: types.CallbackQuery, user: Model, db_manager: DBManager):
    theorys = db_manager.find_data(TheoryModel)
    os.makedirs("./tmp/files/", exist_ok=True)
    filepath = "./tmp/files/teorys.xlsx"
    wb = Workbook()
    if not theorys:
        sheet = wb.active
        
        sheet.cell(1, 1, "1")
        sheet.cell(2, 1, "Название")
        sheet.cell(3, 1, "Описание")
        sheet.cell(4, 1, "Файлы")

        sheet.cell(1, 2, "Теория")
        sheet.cell(2, 2, "Название теориии")
        sheet.cell(3, 2, "Описание теории")
        sheet.cell(4, 2, "Файлы с теорией (jncsdc.wav)")
        sheet.cell(5, 2, "photo2345.png")
        sheet.cell(6, 2, "videoerihvhdbfv.mp4")

        for i in range(3):
            sheet.cell(1, 3+i, f"Параграф {i+1}")
            sheet.cell(2, 3+i, f"Название параграфа {i+1}")
            sheet.cell(3, 3+i, f"Описание параграфа {i+1}")
            def rand_str():
                return str(uuid.uuid4())[:5]
            sheet.cell(4, 3+i, f"Файлы параграфа {i+1} ({rand_str()}.jpg)")
            sheet.cell(5, 3+i, f"Файлы параграфа {i+1} ({rand_str()}.m4a)")
            sheet.cell(6, 3+i, f"Файлы параграфа {i+1} ({rand_str()}.pdf)")
        bot.send_message(user.telegram_id, "Теория отсутствует. Вот пример заполнения таблицы:")
    else:
        for i, theory in enumerate(theorys):
            sheet_name = copy.deepcopy(theory.name)
            sheet_name = "".join(filter(lambda x: x in "йцукенгшщзхфывапролдячсмитьбюЙЦУКЕНГШЩЗХЪъФЫВАПРОЛДЖЭЯЧСМИТЬБЮqwertyuiopasdfghjklzxcvbnmQWERTYUIOPASDFGHJKLXCVBNM",sheet_name))
            if not sheet_name:
                sheet_name = rand_str()
            sheet = wb.create_sheet(sheet_name)
            sheet.cell(1, 1, str(i+1))
            sheet.cell(2, 1, "Название")
            sheet.cell(3, 1, "Описание")
            sheet.cell(4, 1, "Файлы")
            sheet.cell(1, 2, "Теория")
            sheet.cell(2, 2, theory.name)
            sheet.cell(3, 2, theory.description)
            set_files(sheet, theory.files, 2, 4)
            for j, paragraph in enumerate(theory.paragraphs):
                paragraph = replace_none_to_space(paragraph)
                sheet.cell(1, 3+j, str(j+1))
                sheet.cell(2, 3+j, paragraph["name"])
                sheet.cell(3, 3+j, paragraph["description"])
                set_files(sheet, paragraph["files"], 3+j, 4)
        wb.remove(wb[wb.worksheets[0].title])
    wb.save(filepath)
    wb.close()
    bot.send_document(user.telegram_id, open(filepath, "rb"))

def update_file_cb(bot: TeleBot, callback: types.CallbackQuery, user: Model, db_manager: DBManager):
    user.state = "update_file"
    db_manager.save_data(user)
    bot.send_message(user.telegram_id, "Отправьте файл")
    return True

def update_file(bot: TeleBot, filepath: str, message: types.Message, user: Model, db_manager: DBManager):
    user.state = "menu"
    db_manager.save_data(user)
    bot.send_message(user.telegram_id, "Отправьте /menu, чтобы появилась клавиатура")